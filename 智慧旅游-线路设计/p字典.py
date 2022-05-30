
from openpyxl import load_workbook

p={}

excel=load_workbook('D:\\智慧旅游\\智慧旅游-线路设计\\弗洛伊德邻接矩阵-数据版.xlsx')
ws=excel.active

for row in range(2,44):
    row=str(row)
    key=ws['A'+row].value
    p[key]=int(row)-2

print(p)


