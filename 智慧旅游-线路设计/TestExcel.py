#读取Excel中的邻接矩阵信息

import pandas as pd
import xlrd
from openpyxl import load_workbook

#读取工作表中数据
'''
data=pd.read_excel('D://智慧旅游//弗洛伊德邻接矩阵-数据版.xlsx',sheet_name=None)

for i,j in data.items():
    row_data=j.iloc[1:44]
    for a in row_data:
        print(a,end="   ")
    print()
'''

'''
data=pd.read_excel('D:\\智慧旅游\\智慧旅游-线路设计\\弗洛伊德邻接矩阵-数据版.xlsx',sheet_name='Sheet1')
row_data=data.iloc[1:44]
print(type(row_data))
print(row_data)
'''

# #文件地址
# file_lacation='D:\\智慧旅游\\智慧旅游-线路设计\\弗洛伊德邻接矩阵-数据版.xlsx'
#
# #规定矩阵大小
# row_num=col_num=43

#加载表格
#excel=load_workbook(file_lacation)
# ws=excel.active
# #读取单元格
# table=ws.get_sheet_by_name('Sheet1')
#
# #获取单元格值
# data1=table.cell(row=2,column=2).value
# print(data1)


# workbook=xlrd.open_workbook(r'D:\\智慧旅游\\智慧旅游-线路设计\\弗洛伊德邻接矩阵-数据版.xlsx')
#
# sheet2_name=workbook.sheet_names()[1]
# sheet2=workbook.sheet_by_name('Sheet1')
#
# rows=sheet2.row_values(3)
# print(rows)

from openpyxl import load_workbook

excel=load_workbook('D:\\智慧旅游\\智慧旅游-线路设计\\弗洛伊德邻接矩阵-数据版.xlsx')
ws=excel.active
r=[]
m=float('inf')

for row in range(2,44):
    row=str(row)
    r0=[ws['B'+row].value,ws['C'+row].value,ws['D'+row].value,ws['E'+row].value,ws['F'+row].value,ws['G'+row].value,ws['H'+row].value,ws['I'+row].value,ws['J'+row].value,ws['K'+row].value,ws['L'+row].value,ws['M'+row].value,ws['N'+row].value,ws['O'+row].value,ws['P'+row].value,ws['Q'+row].value,ws['R'+row].value,ws['S'+row].value,ws['T'+row].value,ws['U'+row].value,ws['V'+row].value,ws['W'+row].value,ws['X'+row].value,ws['Y'+row].value,ws['Z'+row].value,ws['AA'+row].value,ws['AB'+row].value,ws['AC'+row].value,ws['AD'+row].value,ws['AE'+row].value,ws['AF'+row].value,ws['AG'+row].value,ws['AH'+row].value,ws['AI'+row].value,ws['AJ'+row].value,ws['AK'+row].value,ws['AL'+row].value,ws['AM'+row].value,ws['AN'+row].value,ws['AO'+row].value,ws['AP'+row].value,ws['AQ'+row].value]
    r.append(r0)

for lis in r:
    for item in lis:
        if item=='m':
            item=m
        print(item,end="    ")
    print()


