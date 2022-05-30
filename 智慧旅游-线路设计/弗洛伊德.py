
#邻接矩阵
from openpyxl import load_workbook
import random
import copy
import sys
import math
import tkinter #GUI模块
import threading #多线程模块
from functools import reduce


excel=load_workbook('D:\\智慧旅游\\智慧旅游-线路设计\\弗洛伊德邻接矩阵-数据版.xlsx')
ws=excel.active
r=[]
r_float=[]

m=float('inf')

for row in range(2,44):
    row=str(row)
    r0=[ws['B'+row].value,ws['C'+row].value,ws['D'+row].value,ws['E'+row].value,ws['F'+row].value,ws['G'+row].value,ws['H'+row].value,ws['I'+row].value,ws['J'+row].value,ws['K'+row].value,ws['L'+row].value,ws['M'+row].value,ws['N'+row].value,ws['O'+row].value,ws['P'+row].value,ws['Q'+row].value,ws['R'+row].value,ws['S'+row].value,ws['T'+row].value,ws['U'+row].value,ws['V'+row].value,ws['W'+row].value,ws['X'+row].value,ws['Y'+row].value,ws['Z'+row].value,ws['AA'+row].value,ws['AB'+row].value,ws['AC'+row].value,ws['AD'+row].value,ws['AE'+row].value,ws['AF'+row].value,ws['AG'+row].value,ws['AH'+row].value,ws['AI'+row].value,ws['AJ'+row].value,ws['AK'+row].value,ws['AL'+row].value,ws['AM'+row].value,ws['AN'+row].value,ws['AO'+row].value,ws['AP'+row].value,ws['AQ'+row].value]
    r.append(r0)

for lis in r:
    lis=[m if x=='m' else x for x in lis]
    r_float.append(lis)

'''
for lis in r_float:
    for item in lis:
        print(item,end="    ")
    print()
'''

n=len(r)
# print(n)

#打印邻接矩阵

# for list in r:
#     for item in list:
#         print(item,end="\t")
#     print()

#获得队列矩阵

path=[[0 for i in range(n)] for j in range(n)]

for i in range(n):
    for j in range(n):
        if r_float[i][j]!=0 and r_float[i][j]!=m:
            path[i][j]=j

#计算获得距离矩阵并打印

for x in range(n):
    for i in range(n):
        for j in range(n):
            if r_float[i][x]+r_float[x][j]<r_float[i][j]:
                r_float[i][j]=r_float[i][x]+r_float[x][j]
                path[i][j]=path[i][x]

'''
for i in range(n):
    for j in range(n):
        print(r_float[i][j],end="\t")
    print()
'''

# 此时的r已经是距离矩阵

#为地点A-I赋值

p={}

excel=load_workbook('D:\\智慧旅游\\智慧旅游-线路设计\\弗洛伊德邻接矩阵-数据版.xlsx')
ws=excel.active

for row in range(2,44):
    row=str(row)
    key=ws['A'+row].value
    p[key]=int(row)-2

places=[]
n=0
print("（输入q退出）")
k=input("输入您的起点：")
places.append(k)
while places[n]!='q':
    k=input("输入您想要游览的地点：")
    if k!='q':
        if p.get(k, 0) == 0:
            print("输入地点错误，请重新输入")
        else:
            places.append(k)
            n += 1
    else:
        print("输入完啦，开始计算最优路线")
        break

#print(places)


#计算路程

(ALPHA,BETA,RHO,Q)=(1.0,1.5,0.9,100) #ALPHA：信息启发因子，控制搜索范围 #BETA：期望启发因子，控制收敛速度 #RHO：信息素挥发因子，（1-RHO）表示残留因子
#距离信息和信息素
distance_graph=[[0.0 for col in range(len(places))] for row in range(len(places))]
pheromone_graph=[[1.0 for col in range(len(places))] for row in range(len(places))]

interest_num=len(places)
start=places[0]
N=1


# #整理并打印distance_graph
# for i in range(interest_num):
#     for j in range(interest_num):
#         distance_graph[i][j]=r_float[p[places[i]]][p[places[j]]]
#         print(distance_graph[i][j],end="    ")
#     print()


places.remove(start)
route=[start]
length=0

#起始数据


for i in range(n): #一共比较n轮
    list=[]
    for j in range(len(places)):
        list.append(r_float[p[start]][p[places[j]]])
    length += min(list)
    index_min = list.index(min(list))
    start = places[index_min]
    places.remove(start)
    route.append(start)

print("最短线路长度为%d"%(length))
print("游览顺序为：",end=" ")
for point in route:
    if point!=route[len(route)-1]:
        print(point,end="----->")
    else:
        print(point)

#-----------------------------------------------------------------------------------------------------------------------------------------------------------------

#景点数和距离表达式需要根据实际情况修改
(interest_num,ant_num)=(len(places),50)
#距离信息和信息素（后期需要调整）
distance_graph=[[0.0 for col in range(interest_num)] for row in range(interest_num)]
pheromone_graph=[[1.0 for col in range(interest_num)] for row in range(interest_num)]

#---------------- 蚂蚁 ----------------
class Ant(object):

    # 初始化
    def __init__(self,ID):

        self.ID=ID              #ID
        self.__clean_data()     #随机初始化生成点

    #初始数据
    def __clean_data(self):

        self.path=[]                                                    #当前蚂蚁路径
        self.total_distance=0.0                                         #当前路径总距离
        self.move_count=0                                               #移动次数
        self.current_interest=-1                                        #当前停留的景点
        self.open_table_interest=[True for i in range(interest_num)]        #探索景点的状态

        interest_index=random.randint(0,interest_num-1)                 #随机初始出生点
        self.current_interest=interest_index
        self.path.append(interest_index)
        self.open_table_interest[interest_index]=False                      #遍历
        self.move_count=1

    #选择下一个景点
    def __choice_next_interest(self):
        next_interest=-1
        select_interests_prob=[0.0 for i in range(interest_num)]        #储存去下一个景点的概率
        total_prob=0.0

        #获取去下一个景点的概率
        for i in range(interest_num):
            if self.open_table_interest[i]:
                try:
                    select_interests_prob[i]=pow(pheromone_graph[self.current_interest][i],ALPHA)*pow((1.0/pheromone_graph[self.current_interest][i]),BETA)
                    total_prob+=select_interests_prob[i]
                except ZeroDivisionError as e:
                    print('Ant ID:{ID}, current interest:{current}, target interest:{target}'.format(ID=self.ID,current=self.current_interest,target=i))
                    sys.exit(1)

        #轮盘选择景点
        if total_prob>0.0:
            #产生一个0.0-total_prob之间的随机概率
            temp_prob=random.uniform(0.0,total_prob)
            for i in range(interest_num):
                if self.open_table_interest[i]:
                    #轮次相减
                    temp_prob-=select_interests_prob[i]
                    if temp_prob<0.0:
                        next_interest=i
                        break

        if (next_interest==-1):
            next_interest=random.randint(0,interest_num-1)
            while ((self.open_table_interest[next_interest])==False):       #if==False，说明已经遍历过了
                next_interest=random.randint(0,interest_num-1)

        #返回下一个景点序号
        return next_interest

    #计算路径总距离
    def __cal_total_distance(self):
        temp_distance=0.0
        for i in range(1,interest_num):
            start, end=self.path[i],self.path[i-1]
            temp_distance+=distance_graph[start][end]
        #回路
        end=self.path[0]
        temp_distance+=distance_graph[start][end]
        self.total_distance=temp_distance

    #移动操作
    def __move(self,next_interest):
        self.path.append(next_interest)
        self.open_table_interest[next_interest]=False
        self.total_distance+=distance_graph[self.current_interest][next_interest]
        self.current_interest=next_interest
        self.move_count+=1

    #搜索路径
    def search_path(self):
        #初始化路径
        self.__clean_data()
        #搜索路径，遍历完所有城市为止
        while self.move_count<interest_num:
            #移动到下一景点
            next_interest=self.__choice_next_interest()
            self.__move(next_interest)
        #计算路径总长度
        self.__cal_total_distance()



#---------------- TSP问题 ----------------
class TSP(object):
    def __init__(self,root,width=1000,height=600,n=interest_num):

        #创建画布
        self.root=root
        self.width=width
        self.height=height

        #景点数目初始化为interest_num
        self.n=n

        #tkinter.Canvas
        self.canvas=tkinter.Canvas(
            root,
            width=self.width,
            height=self.height,
            bg="#EBEBEB",#背景白色
            xscrollincrement=1,
            yscrollincrement=1
        )

        self.canvas.pack(expand=tkinter.YES,fill=tkinter.BOTH)
        self.title("TSP蚁群算法(n:初始化 e:开始搜索 s:停止搜索 q:退出程序)")
        self.__r=5
        self.__lock=threading.RLock()       #线程锁
        self.__bindEvents()
        # self.new()

        #计算景点间距离
        for i in range(interest_num):
            for j in range(interest_num):
                distance_graph[i][j]=r_float[p[places[i]]][p[places[j]]]

    #按键响应程序
    def __bindEvents(self):
        self.root.bind("q", self.quite)                      #退出程序
        self.root.bind("n", self.new)                        #初始化
        self.root.bind("e", self.search_path)               #开始搜索
        self.root.bind("s", self.stop)                       #停止搜索

    #更改标题
    def title(self,s):
        self.root.title(s)

    #初始化
    def new(self,evt=None):
        #停止路线
        self.__lock.acquire()
        self.__running=False
        self.__lock.release()
        self.clear()                                        #清楚信息
        self.nodes =[]                                      #节点坐标
        self.nodes2=[]                                      #节点对象

        # #初始化城市节点
        # for i in range(len(distance_x)):
        #     #在画布上随机初始目标
        #     x=distance_x[i]
        #     y=distance_y[i]
        #     self.nodes.append((x,y))
        #     #生成节点椭圆，半径为self.__r
        #     node=self.canvas.create_oval(
        #         x - self.__r,
        #         y - self.__r,
        #         x + self.__r,
        #         y + self.__r,
        #         fill = "#ff0000",
        #         outline = "#000000",
        #         tags = "node",
        #     )
        #     self.nodes2.append(node)
        #     # 显示坐标
        #     self.canvas.create_text(x,y-10,text='('+str(x)+','+str(y)+')',fill='black') #text规定了绘制文字的内容，fill代表绘制文字的颜色

        # #顺序连接景点
        # self.line(range(interest_num))

        #初始景点之间的距离和信息素
        for i in range(interest_num):
            for j in range(interest_num):
                pheromone_graph[i][j]=1.0

        self.ants=[Ant(ID) for ID in range(ant_num)]                #初始蚁群
        self.best_ant=Ant(-1)                                       #初始最优解
        self.best_ant.total_distance=1<<31                          #初始最大距离
        self.iter=1                                                 #初始化迭代次数

    #将节点按order顺序连线
    def line(self,order):
        #删除原线
        self.canvas.delete("line")
        def line2(i1,i2):
            p1,p2=self.nodes[i1],self.nodes[i2]
            self.canvas.create_line(p1,p2,fill="#000000",tags="line")
            return i2
        #order[-1]为初始值
        reduce(line2,order,order[-1])

    #清除画布
    def clear(self):
        for item in self.canvas.find_all():
            self.canvas.delete(item)

    #退出程序
    def quite(self,evt):
        self.__lock.acquire()
        self.__running=False
        self.__lock.release()
        self.root.destroy()
        print(u"\n程序已退出")
        sys.exit()

    #停止搜索
    def stop(self,evt):
        self.__lock.acquire()
        self.__running=False
        self.__lock.release()

    #开始搜索
    def search_path(self,evt=None):
        #开始线程
        self.__lock.acquire()
        self.__running = True
        self.__lock.release()
        while self.__running:
            #遍历每一只蚂蚁
            for ant in self.ants:
                #搜索一条路径
                ant.search_path()
                #与当前最优蚂蚁比较
                if ant.total_distance<self.best_ant.total_distance:
                    #更新最优解
                    self.best_ant=copy.deepcopy(ant)
            #更新信息素
            self.__update_pheromone_graph()
            print(u"迭代次数：",self.iter,u"最佳路径总距离：",int(self.best_ant.total_distance))
            #连线
            self.line(self.best_ant.path)
            #设置标题
            self.title("TSP蚁群算法(n:初始化 e:开始搜索 s:停止搜索 q:退出程序) 迭代次数：%d" % self.iter)
            #更新画布
            self.canvas.update()
            self.iter+=1

    #更新信息素
    def __update_pheromone_graph(self):
        #获取每只蚂蚁在其路径上留下的信息素
        temp_pheromone=[[0.0 for col in range(interest_num)] for row in range(interest_num)]
        for ant in self.ants:
            for i in range(1,interest_num):
                start,end = ant.path[i-1], ant.path[i]
                #在路径上的每两个相邻城市间留下信息素，与路径总距离反比
                temp_pheromone[start][end]+=Q/ant.total_distance
                temp_pheromone[end][start]=temp_pheromone[start][end]
        #更新所有城市之间的信息素，旧信息素衰减加上新迭代信息素
        for i in range(interest_num):
            for j in range(interest_num):
                pheromone_graph[i][j]=pheromone_graph[i][j]*RHO+temp_pheromone[i][j]

    #主循环
    def mainloop(self):
        self.root.mainloop()

#---------------- 程序入口处 ----------------

if __name__ == '__main__':

    TSP(tkinter.Tk()).mainloop()
